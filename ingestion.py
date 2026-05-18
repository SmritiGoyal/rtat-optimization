"""
ingestion.py
============
End-to-end ingestion pipeline: raw Excel sources → integrated master parquet.

Maps directly Steps 1-4:
    - Step 1: Workbook inventory, column profiling, truncation flags
    - Step 2: Streaming join key validation across all domains/years
    - Step 3: Build analysis cohort and OnTime_T targets
    - Step 4A: Reclaim source → repair-level auxiliary feature table
    - Step 4B: Parts ledger → repair-level auxiliary feature table
    - Step 4C: Integrated master table

Produces the following parquet artifacts under ``outputs/interim/``:
    - sheet_inventory.csv         (Step 1)
    - column_profile.csv          (Step 1)
    - truncation_flags.csv        (Step 1)
    - domain_key_stats.csv        (Step 2)
    - pairwise_overlap.csv        (Step 2)
    - cohort_summary.csv          (Step 3)
    - target_dist.csv             (Step 3)
    - ontime_rates.csv            (Step 3)
    - reclaim_features.parquet    (Step 4A)
    - parts_features.parquet      (Step 4B)
    - parts_features_modelsafe.parquet (Step 4B)
    - master_integrated.parquet   (Step 4C)
    - master_train.parquet        (Step 4C — 2023-2025 cohort)
    - master_holdout.parquet      (Step 4C — 2026 cohort)
    - missingness_summary.csv     (Step 4C)
    - data_dictionary.csv         (Step 4C)
"""

from __future__ import annotations

import gc
import logging
import math
import re
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)
logger = logging.getLogger(__name__)


# =====================================================================
# CONFIGURATION
# =====================================================================

PROJECT_ROOT = Path(__file__).resolve().parent


@dataclass(frozen=True)
class IngestionConfig:
    """Configuration for the ingestion pipeline.

    All paths and constants from the original notebook setup cells,
    consolidated into a single frozen dataclass for clarity and safety.
    """
    # Directories
    data_dir: Path = PROJECT_ROOT / "data" / "raw"
    output_dir: Path = PROJECT_ROOT / "outputs" / "interim"

    # Source file patterns (case-insensitive substring match)
    master_pattern: str = "master_repair_data"
    parts_pattern: str = "parts_ledger"
    reclaim_pattern: str = "reclaim_records"

    # Join keys
    master_key: str = "Repair_No"
    parts_keys: tuple[str, ...] = ("Repair_Receipt_No", "Repair_Receipt_No_Merge")
    reclaim_key: str = "GSFS_Repair_Header_No"

    # Target settings
    ontime_thresholds: tuple[int, ...] = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10)
    target_min_days: int = 0
    target_max_days: int = 365
    excel_max_rows: int = 1_048_576
    sample_rows: int = 5_000


# Module-level constants (faithful to notebook)

RECLAIM_PARTS_COLS = ["Parts_No1", "Parts_No2", "Parts_No3", "Parts_No4", "Parts_No5"]

NON_APPLIANCE_DIVISIONS = [
    "HANDSET", "LED Signage", "Signage", "Commercial TV",
    "MNT Signage", "PTV", "Robot Business Task",
]
NON_RTAT_CENTER_TYPES = [
    "Affiliate", "DSC", "COMMERCIAL ASC (Non Referral)",
]

MASTER_COLS = [
    "Repair_No", "Warranty_Closed_Date", "Month_",
    "RTAT_Numerator", "RTAT_Denominator",
    "SVC_Engineer_Code", "SVC_Center_Type", "Channel",
    "Product3_Code", "Product3_Name", "Division_Name",
    "State_", "City_", "General_Market", "Market_Category",
    "Promoter", "Passive", "Detractor",
]

RECLAIM_DESIRED_COLS = [
    "GSFS_Repair_Header_No",
    "Warranty_Flag", "SVC_Center_Type", "SVC_Engineer_Code", "Ship_To_Code",
    "Primary_Defect_Code", "Primary_Defect_Desc",
    "Primary_Repair_Code", "Primary_Repair_Desc",
    "Receipt_Symptom", "SVC_Symptom",
    "Repair_Receipt_Timestamp", "Repair_End_Timestamp",
    "Parts_No1", "Parts_No2", "Parts_No3", "Parts_No4", "Parts_No5",
    "Parts_Desc1", "Parts_Desc2", "Parts_Desc3", "Parts_Desc4", "Parts_Desc5",
    "Reclaim_Period",
    "Same_Symptom_Reclaim", "Same_Servicer_Flag", "Same_Day_Dispatch_Flag",
    "Reversed_Case_Flag", "SVC_Sealed_Repair", "SVC_TER_Repair", "SVC_Part_Usage",
    "Reclaim_Symptom",
    "Division_Code", "Division_Name", "Product2_Name", "Product3_Name", "Model_Code",
]

RECLAIM_FEATURE_COLS = [
    "repair_no_clean", "source_year",
    "Warranty_Flag",
    "Primary_Defect_Code", "Primary_Defect_Desc",
    "Primary_Repair_Code", "Primary_Repair_Desc",
    "Receipt_Symptom", "SVC_Symptom",
    "Division_Code", "Division_Name", "Product2_Name", "Product3_Name", "Model_Code",
    "has_parts_reclaim", "parts_count_reclaim",
    "repair_duration_days", "is_reclaim_case", "reclaim_period_days",
    "same_symptom_reclaim", "same_servicer_flag", "same_day_dispatch_flag",
    "reversed_case_flag", "svc_sealed_repair", "svc_ter_repair", "svc_part_usage_flag",
]

PARTS_KEY = "Repair_Receipt_No"

PARTS_DESIRED_COLS = [
    "Repair_Receipt_No", "Parts_No", "Order_Qty",
    "Order_Timestamp", "Picking_Release_Timestamp",
    "Actual_Shipment_Timestamp", "Arrival_Date",
    "Shipping_Method", "Division_Name", "ProdL2_Name",
    "SO_Type", "SO_Line_Type",
]

PARTS_MODEL_SAFE_COLS = [
    "repair_no_clean",
    "ordered_via_dms",
    "parts_line_count",
    "parts_order_qty_sum",
    "parts_multi_line_flag",
    "parts_has_arrival_flag",
    "parts_has_shipment_flag",
    "parts_shipping_method_first",
    "parts_prodL2_first",
    # parts_so_type_first excluded — 100% constant
    "parts_truncation_flag",
]


# =====================================================================
# SECTION 1: HELPER FUNCTIONS
# =====================================================================

def list_excel_files(folder: Path) -> list[Path]:
    """List .xlsx files in folder, skipping Office lock files (~$ prefix)."""
    return sorted(
        [f for f in folder.iterdir()
         if f.is_file() and f.suffix.lower() == ".xlsx"
         and not f.name.startswith("~$")],
        key=lambda x: x.name.lower()
    )


def classify_domain(name: str, cfg: IngestionConfig) -> str:
    """Map a filename to a source domain (master / parts / reclaim / unknown)."""
    n = name.lower()
    if cfg.master_pattern in n:
        return "master"
    if cfg.parts_pattern in n:
        return "parts"
    if cfg.reclaim_pattern in n:
        return "reclaim"
    return "unknown"


def normalize_col(col) -> str:
    """Strip and collapse whitespace in a column name; return empty for nulls."""
    if col is None or (isinstance(col, float) and np.isnan(col)):
        return ""
    return re.sub(r"\s+", " ", str(col).strip())


def extract_year(text: str):
    """Extract a 4-digit year (20XX) from a string, return NaN if none."""
    m = re.search(r"(20\d{2})", str(text))
    return int(m.group(1)) if m else np.nan


def clean_key_strict(v):
    """Canonical join-key cleaning: uppercase, strip whitespace, drop '.0' suffix.

    Used as the primary key normalizer across master, parts, and reclaim
    tables, which arrived with inconsistent string/numeric formatting.
    Null sentinels (NAN, NONE, NULL, <NA>, empty) are returned as None.
    """
    if v is None:
        return None
    s = re.sub(r"\.0$", "", re.sub(r"\s+", "", str(v).strip().upper()))
    return None if s in {"", "NAN", "NONE", "NULL", "<NA>"} else s


def clean_key_loose(v):
    """Strict cleaning followed by removal of all non-alphanumeric chars.

    Fallback for keys with embedded separators that escape strict matching.
    """
    s = clean_key_strict(v)
    if s is None:
        return None
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s if s else None


# Alias preserved for back-compat with notebook code blocks that used
# the shorter name in Steps 4A/4B/4C.
clean_key = clean_key_strict


def build_header_index(header_row) -> dict:
    """Build {lowercased_normalized_name: column_index} for a header row."""
    return {normalize_col(c).lower(): i
            for i, c in enumerate(header_row)
            if normalize_col(c)}


# Notebook alias used in Step 4B
build_header_idx = build_header_index


def to_numeric_safe(s: pd.Series) -> pd.Series:
    """Coerce to numeric; non-numeric values become NaN."""
    return pd.to_numeric(s, errors="coerce")


def to_datetime_safe(s: pd.Series) -> pd.Series:
    """Coerce to datetime; un-parseable values become NaT."""
    return pd.to_datetime(s, errors="coerce")


def normalize_flag(v):
    """Parse a binary flag from heterogeneous reclaim-record formats.

    Positive values: Y, YES, 1, TRUE, SEALED REPAIR, PART USED, X
    Negative values: N, NO, 0, FALSE
    Null sentinels: NaN, None, empty → pd.NA (preserved for reclaim-only fields)
    """
    if v is None:
        return pd.NA
    s = str(v).strip().upper()
    if s in {"", "NAN", "NONE", "NULL", "<NA>", "NAT"}:
        return pd.NA
    if s in {"Y", "YES", "1", "TRUE", "SEALED REPAIR", "PART USED", "X"}:
        return 1
    if s in {"N", "NO", "0", "FALSE"}:
        return 0
    return pd.NA


def sanitize_for_parquet(df: pd.DataFrame) -> pd.DataFrame:
    """Cast mixed-type object columns to string so PyArrow doesn't choke.

    Object columns sometimes contain literal "nan", "None", "<NA>", or
    "NaT" strings that survived an upstream cast. Normalizes these to
    proper pd.NA before parquet serialization.
    """
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = (out[col].astype(str)
                        .replace({"None": "", "nan": "", "<NA>": "", "NaT": ""})
                        .replace("", pd.NA))
    return out


def simplify_shipping(method) -> str:
    """Bucket raw shipping method strings into 5 canonical tiers + UNKNOWN."""
    if pd.isna(method) or str(method).strip() == "":
        return "UNKNOWN"
    m = str(method).upper()
    if "OVERNIGHT" in m or "NEXT DAY" in m:
        return "OVERNIGHT"
    if "2ND DAY" in m or "2DAY" in m:
        return "TWO_DAY"
    if "GROUND" in m:
        return "GROUND"
    if "PICKUP" in m:
        return "PICKUP"
    return "OTHER"


def is_key_col(col: str) -> bool:
    """Heuristic: does a column name look like a join key?"""
    c = col.lower()
    return any(p in c for p in [
        "repair_no", "repair", "receipt_no", "receipt", "header_no",
        "claim_no", "case_no", "serial_no", "engineer_code", "ship_to_code",
        "parts_no"
    ])


def is_date_col(col: str) -> bool:
    """Heuristic: does a column name look like a date/timestamp?"""
    c = col.lower()
    return any(p in c for p in [
        "date", "timestamp", "time", "closed", "receipt", "arrival",
        "ship", "order", "picking", "created", "updated", "visit"
    ])


def classify_col_type(s: pd.Series) -> str:
    """Heuristically classify a column as date/numeric/text/all_null.

    Used in Step 1 column profiling to flag dates and numeric-like text
    fields that need explicit parsing downstream.
    """
    non_null = s.dropna()
    if len(non_null) == 0:
        return "all_null"
    if pd.to_datetime(non_null, errors="coerce").notna().mean() >= 0.8:
        return "date_like"
    if pd.to_numeric(non_null.astype(str).str.replace(",", "", regex=False),
                     errors="coerce").notna().mean() >= 0.8:
        return "numeric_like"
    return "text_like"


def read_cols(fp: Path, sheet: str, wanted: list[str]) -> pd.DataFrame:
    """Read only the requested columns from a sheet, case-insensitive name match.

    Returns a DataFrame with exactly the columns in ``wanted``, in that
    order. Columns not present in the source are filled with pd.NA.
    """
    header = pd.read_excel(fp, sheet_name=sheet, nrows=0,
                           dtype=object, engine="openpyxl")
    lookup = {normalize_col(c).lower(): normalize_col(c) for c in header.columns}
    usecols = [lookup[w.lower()] for w in wanted if w.lower() in lookup]
    if not usecols:
        return pd.DataFrame()
    df = pd.read_excel(fp, sheet_name=sheet, usecols=usecols,
                       dtype=object, engine="openpyxl")
    df.columns = [normalize_col(c) for c in df.columns]
    rename = {normalize_col(lookup[w.lower()]): w
              for w in wanted if w.lower() in lookup}
    df = df.rename(columns=rename)
    for w in wanted:
        if w not in df.columns:
            df[w] = pd.NA
    return df[wanted]


def list_reclaim_files(folder: Path, cfg: IngestionConfig) -> list[Path]:
    """Return reclaim Excel files in folder, excluding Office lock files."""
    return sorted([
        f for f in folder.glob("*.xlsx")
        if cfg.reclaim_pattern in f.name.lower() and not f.name.startswith("~$")
    ], key=lambda x: x.name.lower())


# =====================================================================
# SECTION 2: STEP 1 — WORKBOOK INVENTORY & COLUMN PROFILING
# =====================================================================

def run_step1_inventory(cfg: IngestionConfig) -> dict[str, pd.DataFrame]:
    """Build sheet inventory, column profile, and truncation-flag tables.

    For every Excel sheet in ``cfg.data_dir``, records the row/column count
    and whether the sheet hit Excel's 1,048,576-row limit (truncated). For
    each sheet, samples ``cfg.sample_rows`` rows and profiles each column's
    null rate, cardinality, and inferred type.

    Returns:
        Dict containing ``sheet_inventory``, ``column_profile``, and
        ``truncation_flags`` DataFrames. All three are also written as
        CSV files under ``cfg.output_dir``.
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    excel_files = list_excel_files(cfg.data_dir)

    # --- 1A: Sheet inventory ---
    sheet_rows = []
    for fp in excel_files:
        try:
            wb = load_workbook(fp, read_only=True, data_only=True)
            for ws in wb.worksheets:
                sheet_rows.append({
                    "file_name": fp.name,
                    "domain": classify_domain(fp.name, cfg),
                    "sheet_name": ws.title,
                    "sheet_max_row": ws.max_row,
                    "sheet_max_col": ws.max_column,
                    "row_limit_hit": ws.max_row == cfg.excel_max_rows,
                })
            wb.close()
        except Exception as e:
            sheet_rows.append({"file_name": fp.name, "error": str(e)})

    sheet_inventory = pd.DataFrame(sheet_rows)

    # --- 1B: Column profiling (sample-based) ---
    col_rows = []
    for _, sr in sheet_inventory.dropna(subset=["sheet_name"]).iterrows():
        fp = cfg.data_dir / sr["file_name"]
        try:
            df = pd.read_excel(fp, sheet_name=sr["sheet_name"],
                               nrows=cfg.sample_rows, dtype=object, engine="openpyxl")
            df.columns = [normalize_col(c) for c in df.columns]
            for col in df.columns:
                s = df[col]
                nn = int(s.notna().sum())
                uq = int(s.dropna().astype(str).nunique())
                col_rows.append({
                    "file_name": sr["file_name"],
                    "sheet_name": sr["sheet_name"],
                    "column_name": col,
                    "non_null_pct": round(nn / len(df), 4) if len(df) else np.nan,
                    "unique_count": uq,
                    "unique_ratio": round(uq / nn, 4) if nn else np.nan,
                    "type_guess": classify_col_type(s),
                    "is_key_col": is_key_col(col),
                    "is_date_col": is_date_col(col),
                })
        except Exception as e:
            col_rows.append({"file_name": sr["file_name"],
                             "sheet_name": sr["sheet_name"],
                             "error": str(e)})

    column_profile = pd.DataFrame(col_rows)
    truncation_flags = sheet_inventory[sheet_inventory["row_limit_hit"] == True]

    # Save
    sheet_inventory.to_csv(cfg.output_dir / "sheet_inventory.csv", index=False)
    column_profile.to_csv(cfg.output_dir / "column_profile.csv", index=False)
    truncation_flags.to_csv(cfg.output_dir / "truncation_flags.csv", index=False)

    logger.info("STEP 1 — sheets: %d, columns profiled: %d, truncated: %d",
                len(sheet_inventory), len(column_profile), len(truncation_flags))

    return {
        "sheet_inventory": sheet_inventory,
        "column_profile": column_profile,
        "truncation_flags": truncation_flags,
    }


# =====================================================================
# SECTION 3: STEP 2 — STREAMING JOIN-KEY VALIDATION
# =====================================================================

def _make_key_state() -> dict:
    """Initialize a fresh key-counter state for one (domain, year, key) bucket."""
    return {
        "row_count": 0, "non_null": 0,
        "seen_once": set(), "seen_multi": set(),
        "dup_rows": 0,
    }


def _update_state(state: dict, raw) -> None:
    """Update a key-counter state with one raw key value."""
    state["row_count"] += 1
    v = clean_key_strict(raw)
    if v is None:
        return
    state["non_null"] += 1
    if v not in state["seen_once"]:
        state["seen_once"].add(v)
    elif v not in state["seen_multi"]:
        state["seen_multi"].add(v)
        state["dup_rows"] += 2
    else:
        state["dup_rows"] += 1


def _state_summary(domain: str, year, key_col: str, state: dict) -> dict:
    """Reduce a key-counter state into a flat dict summary."""
    nn = state["non_null"]
    uq = len(state["seen_once"])
    return {
        "domain": domain,
        "year": year,
        "key_column": key_col,
        "row_count": state["row_count"],
        "non_null_count": nn,
        "non_null_pct": round(nn / state["row_count"], 6) if state["row_count"] else np.nan,
        "unique_count": uq,
        "unique_ratio": round(uq / nn, 6) if nn else np.nan,
        "dup_rows": state["dup_rows"],
    }


def _overlap_stats(left, right, year, l_src, l_key, r_src, r_key) -> dict:
    """Compute intersection / match / Jaccard between two key sets."""
    inter = left & right
    union = left | right
    return {
        "year": year,
        "left_source": l_src, "left_key": l_key,
        "right_source": r_src, "right_key": r_key,
        "left_unique": len(left),
        "right_unique": len(right),
        "intersection": len(inter),
        "left_match_rate": round(len(inter) / len(left), 6) if left else np.nan,
        "right_match_rate": round(len(inter) / len(right), 6) if right else np.nan,
        "jaccard": round(len(inter) / len(union), 6) if union else np.nan,
    }


def run_step2_key_validation(cfg: IngestionConfig) -> dict[str, pd.DataFrame]:
    """Stream every sheet to compute per-domain key statistics + overlaps.

    Memory-efficient: never loads a sheet into RAM — iterates row by row
    using openpyxl's read-only workbook mode and accumulates per-key
    state in sets. Produces:
        - Per-(domain, year, key) row count, null rate, uniqueness
        - Pairwise overlap (master↔parts, master↔reclaim) per year and ALL
        - Parts-internal comparison of Repair_Receipt_No vs ..._Merge
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    dms_compare = defaultdict(lambda: {"rows": 0, "both": 0, "equal": 0})
    key_states: dict[tuple, dict] = {}

    for fp in list_excel_files(cfg.data_dir):
        domain = classify_domain(fp.name, cfg)
        if domain == "unknown":
            continue

        wb = load_workbook(fp, read_only=True, data_only=True)
        try:
            for sname in wb.sheetnames:
                year = extract_year(sname)
                ws = wb[sname]
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue
                hidx = build_header_index(header)

                if domain == "master":
                    needed = [cfg.master_key]
                elif domain == "parts":
                    needed = list(cfg.parts_keys)
                else:
                    needed = [cfg.reclaim_key]

                avail = {k: hidx.get(k.lower()) for k in needed}
                avail = {k: v for k, v in avail.items() if v is not None}
                if not avail:
                    continue

                for k in avail:
                    key_states.setdefault((domain, year, k), _make_key_state())

                for row in rows:
                    if domain == "master":
                        _update_state(
                            key_states[(domain, year, cfg.master_key)],
                            row[avail[cfg.master_key]],
                        )
                    elif domain == "parts":
                        raw0 = row[avail[cfg.parts_keys[0]]] if cfg.parts_keys[0] in avail else None
                        raw1 = row[avail[cfg.parts_keys[1]]] if cfg.parts_keys[1] in avail else None
                        if cfg.parts_keys[0] in avail:
                            _update_state(key_states[(domain, year, cfg.parts_keys[0])], raw0)
                        if cfg.parts_keys[1] in avail:
                            _update_state(key_states[(domain, year, cfg.parts_keys[1])], raw1)
                        # Parts-internal comparison
                        v0 = clean_key_strict(raw0)
                        v1 = clean_key_strict(raw1)
                        dc = dms_compare[year]
                        dc["rows"] += 1
                        if v0 and v1:
                            dc["both"] += 1
                            if v0 == v1:
                                dc["equal"] += 1
                    else:
                        _update_state(
                            key_states[(domain, year, cfg.reclaim_key)],
                            row[avail[cfg.reclaim_key]],
                        )
                gc.collect()
        finally:
            wb.close()
            gc.collect()

    # Build domain-key stats
    domain_key_stats = pd.DataFrame([
        _state_summary(d, y, k, state)
        for (d, y, k), state in sorted(key_states.items())
    ])

    # Pairwise overlap
    def get_keys(domain: str, key_col: str) -> set:
        """Union all seen-once key sets across years for one (domain, key_col)."""
        merged: set = set()
        for (d, y, k), state in key_states.items():
            if d == domain and k == key_col:
                merged.update(state["seen_once"])
        return merged

    overlap_rows = []
    years = sorted(set(y for (_, y, _) in key_states if isinstance(y, int)))

    for year in years:
        m = key_states.get(("master", year, cfg.master_key))
        p0 = key_states.get(("parts", year, cfg.parts_keys[0]))
        r = key_states.get(("reclaim", year, cfg.reclaim_key))

        if m and p0:
            overlap_rows.append(_overlap_stats(
                m["seen_once"], p0["seen_once"], year,
                "master", cfg.master_key, "parts", cfg.parts_keys[0]))
        if m and r:
            overlap_rows.append(_overlap_stats(
                m["seen_once"], r["seen_once"], year,
                "master", cfg.master_key, "reclaim", cfg.reclaim_key))

    # ALL-years overlap
    overlap_rows.append(_overlap_stats(
        get_keys("master", cfg.master_key),
        get_keys("parts", cfg.parts_keys[0]),
        "ALL", "master", cfg.master_key, "parts", cfg.parts_keys[0]))
    overlap_rows.append(_overlap_stats(
        get_keys("master", cfg.master_key),
        get_keys("reclaim", cfg.reclaim_key),
        "ALL", "master", cfg.master_key, "reclaim", cfg.reclaim_key))

    pairwise_overlap = pd.DataFrame(overlap_rows)

    # Parts-internal comparison
    dms_compare_df = pd.DataFrame([
        {"year": y, "rows": s["rows"], "both_non_null": s["both"],
         "equal_rows": s["equal"],
         "equal_rate": round(s["equal"] / s["both"], 6) if s["both"] else np.nan}
        for y, s in sorted(dms_compare.items())
    ])

    # Save
    domain_key_stats.to_csv(cfg.output_dir / "domain_key_stats.csv", index=False)
    pairwise_overlap.to_csv(cfg.output_dir / "pairwise_overlap.csv", index=False)
    dms_compare_df.to_csv(cfg.output_dir / "parts_internal_compare.csv", index=False)

    logger.info("STEP 2 — key buckets: %d, overlap rows: %d",
                len(domain_key_stats), len(pairwise_overlap))

    return {
        "domain_key_stats": domain_key_stats,
        "pairwise_overlap": pairwise_overlap,
        "parts_internal_compare": dms_compare_df,
    }


# =====================================================================
# SECTION 4: STEP 3 — COHORT BUILD + OnTime_T TARGETS
# =====================================================================

DESIRED_MASTER_COLS_STEP3 = [
    "Repair_No", "Warranty_Closed_Date", "Month_",
    "RTAT_Numerator", "RTAT_Denominator",
    "SVC_Engineer_Code", "SVC_Center_Type", "Channel",
    "Product3_Code", "Product3_Name", "Division_Name",
    "State_", "City_", "General_Market", "Market_Category",
    "Promoter", "Passive", "Detractor",
]


def _add_targets(df: pd.DataFrame, year, cfg: IngestionConfig) -> pd.DataFrame:
    """Add cohort flags, target_days, and OnTime_T columns to a master frame.

    Cohort flags (all four required for inclusion):
        flag_valid_id    — repair_no_clean is not null
        flag_valid_rtat  — RTAT_Numerator parses as numeric
        flag_nonneg      — RTAT >= TARGET_MIN_DAYS (0)
        flag_within_max  — RTAT <= TARGET_MAX_DAYS (365)

    Target columns:
        target_days       — RTAT for cohort rows, NaN otherwise
        OnTime_{T} for T in cfg.ontime_thresholds — binary per threshold
        has_nps           — derived from Promoter/Passive/Detractor sum
    """
    df = df.copy()
    df["source_year"] = year
    df["repair_no_clean"] = (df["Repair_No"].astype("string")
                             .str.strip().str.upper()
                             .str.replace(r"\.0$", "", regex=True)
                             .str.replace(r"\s+", "", regex=True)
                             .replace({"": pd.NA, "NAN": pd.NA, "NONE": pd.NA}))
    df["rtat_days"] = to_numeric_safe(df["RTAT_Numerator"]).astype("float32")
    df["rtat_denominator"] = to_numeric_safe(df["RTAT_Denominator"]).astype("float32")
    df["closed_dt"] = to_datetime_safe(df["Warranty_Closed_Date"])

    df["flag_valid_id"] = df["repair_no_clean"].notna()
    df["flag_valid_rtat"] = df["rtat_days"].notna()
    df["flag_nonneg"] = df["rtat_days"] >= cfg.target_min_days
    df["flag_within_max"] = df["rtat_days"] <= cfg.target_max_days
    df["flag_cohort"] = (df["flag_valid_id"] & df["flag_valid_rtat"]
                         & df["flag_nonneg"] & df["flag_within_max"])

    df["target_days"] = np.where(
        df["flag_cohort"], df["rtat_days"], np.nan
    ).astype("float32")

    for t in cfg.ontime_thresholds:
        df[f"OnTime_{t}"] = pd.array(
            np.where(df["flag_cohort"],
                     (df["target_days"] <= t).astype("int8"), pd.NA),
            dtype="Int8"
        )

    nps_sum = (to_numeric_safe(df["Promoter"]).fillna(0)
               + to_numeric_safe(df["Passive"]).fillna(0)
               + to_numeric_safe(df["Detractor"]).fillna(0))
    df["has_nps"] = nps_sum > 0
    return df


def _cohort_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Per-year + ALL summary of cohort sizes, missing RTAT, NPS coverage."""
    rows = []
    for yr, g in list(df.groupby("source_year")) + [("ALL", df)]:
        rows.append({
            "year": yr,
            "rows_total": len(g),
            "rows_cohort": int(g["flag_cohort"].sum()),
            "cohort_rate": round(float(g["flag_cohort"].mean()), 4),
            "rows_missing_rtat": int(g["flag_valid_rtat"].eq(False).sum()),
            "rows_negative_rtat": int((g["rtat_days"] < 0).sum()),
            "rows_with_nps": int(g["has_nps"].sum()),
            "nps_rate": round(float(g["has_nps"].mean()), 4),
        })
    return pd.DataFrame(rows)


def _target_dist(df: pd.DataFrame) -> pd.DataFrame:
    """Per-year + ALL distributional stats on target_days (cohort rows only)."""
    cohort = df[df["flag_cohort"]]
    rows = []
    for yr, g in list(cohort.groupby("source_year")) + [("ALL", cohort)]:
        rows.append({
            "year": yr, "n": len(g),
            "mean": round(float(g["target_days"].mean()), 3),
            "median": round(float(g["target_days"].median()), 3),
            "std": round(float(g["target_days"].std()), 3),
            "p25": round(float(g["target_days"].quantile(.25)), 3),
            "p75": round(float(g["target_days"].quantile(.75)), 3),
            "p90": round(float(g["target_days"].quantile(.90)), 3),
            "p95": round(float(g["target_days"].quantile(.95)), 3),
            "min": round(float(g["target_days"].min()), 3),
            "max": round(float(g["target_days"].max()), 3),
        })
    return pd.DataFrame(rows)


def _ontime_rates(df: pd.DataFrame, cfg: IngestionConfig) -> pd.DataFrame:
    """Per-year + ALL on-time rate at every threshold T in cfg.ontime_thresholds."""
    cohort = df[df["flag_cohort"]]
    rows = []
    for yr, g in list(cohort.groupby("source_year")) + [("ALL", cohort)]:
        row = {"year": yr, "n": len(g)}
        for t in cfg.ontime_thresholds:
            row[f"OnTime_{t}"] = round(float(g[f"OnTime_{t}"].astype("float").mean()), 4)
        rows.append(row)
    return pd.DataFrame(rows)


def run_step3_cohort(cfg: IngestionConfig) -> dict[str, pd.DataFrame]:
    """Load master file across all sheets and produce the cohort + targets.

    This is the first-pass cohort build using only the master table; it
    does not yet incorporate exclusions from divisions or center types
    (those come in Step 4C). Step 3 is retained as-is to preserve the
    original notebook's diagnostic outputs (cohort/dist/ontime).
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    master_file = next(
        (f for f in list_excel_files(cfg.data_dir)
         if cfg.master_pattern in f.name.lower()), None
    )
    assert master_file, f"Master file not found in {cfg.data_dir} (pattern: {cfg.master_pattern})"
    logger.info("STEP 3 — master file: %s", master_file.name)

    xls = pd.ExcelFile(master_file, engine="openpyxl")
    frames = []
    for sheet in xls.sheet_names:
        year = extract_year(sheet)
        df = read_cols(master_file, sheet, DESIRED_MASTER_COLS_STEP3)
        if df.empty:
            continue
        df = _add_targets(df, year, cfg)
        frames.append(df)
        gc.collect()

    master = pd.concat(frames, ignore_index=True)
    del frames
    gc.collect()

    cohort_df = _cohort_summary(master)
    dist_df = _target_dist(master)
    ontime_df = _ontime_rates(master, cfg)

    cohort_df.to_csv(cfg.output_dir / "cohort_summary.csv", index=False)
    dist_df.to_csv(cfg.output_dir / "target_dist.csv", index=False)
    ontime_df.to_csv(cfg.output_dir / "ontime_rates.csv", index=False)

    logger.info("STEP 3 — total rows: %s, cohort rows: %s",
                f"{len(master):,}", f"{int(master['flag_cohort'].sum()):,}")

    return {
        "cohort_summary": cohort_df,
        "target_dist": dist_df,
        "ontime_rates": ontime_df,
    }


# =====================================================================
# SECTION 5: STEP 4A — RECLAIM FEATURES
# =====================================================================

def run_step4a_reclaim(cfg: IngestionConfig) -> pd.DataFrame:
    """Build the repair-level reclaim feature table.

    Reads every reclaim Excel (one per year), normalizes column names,
    parses the binary flag columns (Same_Symptom, Sealed, TER, etc.),
    derives ``has_parts_reclaim`` and ``parts_count_reclaim`` from the
    five Parts_No columns, and computes repair duration in days from
    the receipt and end timestamps. Dedups by repair_no_clean keeping
    the earliest year's record.

    Returns and persists the reclaim feature table as parquet.
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    reclaim_files = list_reclaim_files(cfg.data_dir, cfg)

    # Duplicate-file guard preserved from original
    file_names = [f.name.lower() for f in reclaim_files]
    if "reclaim.xlsx" in file_names and "2023 reclaim.xlsx" in file_names:
        reclaim_files = [f for f in reclaim_files if f.name.lower() != "reclaim.xlsx"]
        logger.warning("Duplicate 2023 Reclaim detected — excluded 'Reclaim.xlsx'")

    logger.info("STEP 4A — processing %d reclaim file(s)", len(reclaim_files))
    reclaim_frames = []

    for fp in reclaim_files:
        xls = pd.ExcelFile(fp, engine="openpyxl")
        for sheet in xls.sheet_names:
            year = int(m.group(1)) if (m := re.search(r"(20\d{2})", sheet)) else None
            df = read_cols(fp, sheet, RECLAIM_DESIRED_COLS)
            if df.empty:
                continue

            df["source_year"] = year
            df["reclaim_source"] = fp.name

            # Clean join key
            df["repair_no_clean"] = df[cfg.reclaim_key].apply(clean_key)

            # True parts-required flag from embedded part columns
            parts_present = df[RECLAIM_PARTS_COLS].apply(
                lambda col: col.astype(str).str.strip()
                            .replace({"nan": "", "None": "", "<NA>": ""})
                            .ne("")
            )
            df["has_parts_reclaim"] = parts_present.any(axis=1).astype("Int8")
            df["parts_count_reclaim"] = parts_present.sum(axis=1).astype("Int8")

            # Duration from reclaim timestamps
            df["repair_receipt_dt"] = to_datetime_safe(df["Repair_Receipt_Timestamp"])
            df["repair_end_dt"] = to_datetime_safe(df["Repair_End_Timestamp"])
            df["repair_duration_days"] = (
                (df["repair_end_dt"] - df["repair_receipt_dt"])
                .dt.total_seconds() / 86400
            ).astype("float32")

            # Reclaim period
            df["reclaim_period_days"] = to_numeric_safe(df["Reclaim_Period"]).astype("float32")
            df["is_reclaim_case"] = (
                df["reclaim_period_days"].notna()
                & (df["reclaim_period_days"] >= 0)
            ).astype("Int8")

            # Binary flags
            flag_map = {
                "Same_Symptom_Reclaim": "same_symptom_reclaim",
                "Same_Servicer_Flag": "same_servicer_flag",
                "Same_Day_Dispatch_Flag": "same_day_dispatch_flag",
                "Reversed_Case_Flag": "reversed_case_flag",
                "SVC_Sealed_Repair": "svc_sealed_repair",
                "SVC_TER_Repair": "svc_ter_repair",
                "SVC_Part_Usage": "svc_part_usage_flag",
            }
            for src_col, tgt_col in flag_map.items():
                df[tgt_col] = df[src_col].apply(normalize_flag).astype("Int8")

            reclaim_frames.append(df)
            logger.info(
                "  %s / %s: %s rows, has_parts=%s (%.1f%%), sealed=%s, ter=%s",
                fp.name, sheet,
                f"{len(df):,}",
                f"{int(df['has_parts_reclaim'].sum()):,}",
                100 * df['has_parts_reclaim'].mean(),
                f"{int(df['svc_sealed_repair'].sum()):,}",
                f"{int(df['svc_ter_repair'].sum()):,}",
            )
            gc.collect()

    reclaim_raw = pd.concat(reclaim_frames, ignore_index=True)
    del reclaim_frames
    gc.collect()

    logger.info("STEP 4A — pre-dedup: %s rows", f"{len(reclaim_raw):,}")
    reclaim_raw = reclaim_raw.sort_values(
        ["repair_no_clean", "source_year"], na_position="last"
    ).drop_duplicates(subset="repair_no_clean", keep="first")
    logger.info("STEP 4A — post-dedup: %s rows", f"{len(reclaim_raw):,}")

    reclaim_features = reclaim_raw[
        [c for c in RECLAIM_FEATURE_COLS if c in reclaim_raw.columns]
    ].copy()

    reclaim_clean = sanitize_for_parquet(reclaim_features)
    reclaim_clean.to_parquet(cfg.output_dir / "reclaim_features.parquet", index=False)
    logger.info("STEP 4A — saved reclaim_features.parquet, shape: %s", reclaim_clean.shape)

    return reclaim_clean


# =====================================================================
# SECTION 6: STEP 4B — PARTS LEDGER FEATURES
# =====================================================================

def _norm_text(v):
    """Strip+upper text, return None for null sentinels."""
    if v is None:
        return None
    s = str(v).strip().upper()
    return None if s in {"", "NAN", "NONE", "NULL", "<NA>"} else s


def _to_float(v):
    """Parse to float, return None if not parseable (handles thousands sep)."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", ""))
        except Exception:
            return None


def _to_ts(v):
    """Parse to pd.Timestamp, return None on failure or NaT."""
    if v is None or v == "":
        return None
    if isinstance(v, pd.Timestamp):
        return v
    if isinstance(v, datetime):
        return pd.Timestamp(v)
    try:
        ts = pd.Timestamp(v)
        return None if pd.isna(ts) else ts
    except Exception:
        return None


def _update_min(cur, new):
    """Return min(cur, new) treating None as missing."""
    if new is None:
        return cur
    return new if cur is None else min(cur, new)


def _update_max(cur, new):
    """Return max(cur, new) treating None as missing."""
    if new is None:
        return cur
    return new if cur is None else max(cur, new)


def _first_non_null(s: pd.Series):
    """Return the first non-null value in a Series, or pd.NA if all null."""
    v = s.dropna()
    return v.iloc[0] if len(v) else pd.NA


def _init_parts_state(truncated: bool) -> dict:
    """Initialize a fresh per-repair aggregation state for parts streaming."""
    return {
        "line_count": 0,
        "order_qty_sum": 0.0,
        "parts_no_count": 0,
        "order_ts_min": None,
        "pick_ts_min": None,
        "ship_ts_min": None,
        "arrival_dt_min": None,
        "arrival_dt_max": None,
        "ship_count": 0,
        "arrival_count": 0,
        "shipping_method_first": None,
        "prodL2_first": None,
        "so_type_first": None,
        "truncation_flag": int(truncated),
    }


def _days_between_cols(df: pd.DataFrame, a: str, b: str) -> pd.Series:
    """Compute (b - a) in days as float32, parsing both columns as datetimes."""
    a_ = pd.to_datetime(df[a], errors="coerce")
    b_ = pd.to_datetime(df[b], errors="coerce")
    return ((b_ - a_).dt.total_seconds() / 86400).astype("float32")


def _clean_leg_durations(df: pd.DataFrame) -> pd.DataFrame:
    """Null out negative or implausibly long leg durations (>90 days).

    Negative values are data-entry errors (arrival before ship, etc.).
    Values above 90 days are administrative artifacts (e.g., backordered
    parts kept open in the system) and would bias delivery-tier features.
    """
    leg_cols = [
        "parts_order_to_pick_days_eda",
        "parts_pick_to_ship_days_eda",
        "parts_ship_to_arrival_days_eda",
        "parts_order_to_arrival_days_eda",
    ]
    for col in leg_cols:
        if col in df.columns:
            df[col] = df[col].where(
                df[col].between(0, 90), other=np.nan
            ).astype("float32")
    return df


def run_step4b_parts(cfg: IngestionConfig) -> pd.DataFrame:
    """Build the repair-level parts feature table from the parts ledger.

    Streams every parts Excel sheet row by row (using openpyxl read-only)
    to avoid loading the multi-million row sheets into memory. For each
    repair, accumulates:
        - Line count, total order quantity, count of distinct parts
        - Min/max timestamps for order, picking, shipment, arrival
        - First-seen shipping method, ProdL2, SO type
        - Truncation flag if the source sheet hit Excel's row limit

    Then concatenates per-sheet outputs and re-aggregates across years
    (a repair may have parts ordered in multiple years), computes flag
    columns, derives leg durations, and cleans implausible durations.

    Saves both the full features and the model-safe subset as parquet.
    Returns the full features DataFrame.
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    parts_files = sorted([
        f for f in cfg.data_dir.glob("*.xlsx")
        if cfg.parts_pattern in f.name.lower() and not f.name.startswith("~$")
    ])
    logger.info("STEP 4B — parts files: %d", len(parts_files))

    yearly_paths: list[Path] = []

    for fp in parts_files:
        wb = load_workbook(fp, read_only=True, data_only=True)
        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                row_iter = ws.iter_rows(values_only=True)
                try:
                    raw_header = next(row_iter)
                except StopIteration:
                    continue

                header = [normalize_col(c) for c in raw_header]
                hidx = build_header_idx(header)
                year_m = re.search(r"(20\d{2})", sname)
                year_val = int(year_m.group(1)) if year_m else None

                col_idx: dict[str, int] = {}
                for col in PARTS_DESIRED_COLS:
                    hit = hidx.get(col.lower())
                    if hit is not None:
                        col_idx[col] = hit

                if PARTS_KEY not in col_idx:
                    logger.warning("  Skipped (no key): %s / %s", fp.name, sname)
                    continue

                aggs: dict[str, dict] = {}
                total = 0

                for row in row_iter:
                    total += 1
                    rnn = clean_key(row[col_idx[PARTS_KEY]])
                    if rnn is None:
                        continue

                    truncated = (total + 1) >= cfg.excel_max_rows
                    if rnn not in aggs:
                        aggs[rnn] = _init_parts_state(truncated)
                    state = aggs[rnn]

                    state["line_count"] += 1

                    if "Order_Qty" in col_idx:
                        qty = _to_float(row[col_idx["Order_Qty"]])
                        if qty and not math.isnan(qty):
                            state["order_qty_sum"] += qty

                    if "Parts_No" in col_idx:
                        if clean_key(row[col_idx["Parts_No"]]) is not None:
                            state["parts_no_count"] += 1

                    order_ts = None
                    if "Order_Timestamp" in col_idx:
                        order_ts = _to_ts(row[col_idx["Order_Timestamp"]])
                    state["order_ts_min"] = _update_min(state["order_ts_min"], order_ts)

                    if "Picking_Release_Timestamp" in col_idx:
                        pick_ts = _to_ts(row[col_idx["Picking_Release_Timestamp"]])
                        state["pick_ts_min"] = _update_min(state["pick_ts_min"], pick_ts)

                    if "Actual_Shipment_Timestamp" in col_idx:
                        ship_ts = _to_ts(row[col_idx["Actual_Shipment_Timestamp"]])
                        if ship_ts:
                            state["ship_count"] += 1
                            state["ship_ts_min"] = _update_min(state["ship_ts_min"], ship_ts)

                    if "Arrival_Date" in col_idx:
                        arr_dt = _to_ts(row[col_idx["Arrival_Date"]])
                        if arr_dt:
                            state["arrival_count"] += 1
                            state["arrival_dt_min"] = _update_min(state["arrival_dt_min"], arr_dt)
                            state["arrival_dt_max"] = _update_max(state["arrival_dt_max"], arr_dt)

                    if "Shipping_Method" in col_idx and state["shipping_method_first"] is None:
                        state["shipping_method_first"] = _norm_text(row[col_idx["Shipping_Method"]])
                    if "ProdL2_Name" in col_idx and state["prodL2_first"] is None:
                        state["prodL2_first"] = _norm_text(row[col_idx["ProdL2_Name"]])
                    if "SO_Type" in col_idx and state["so_type_first"] is None:
                        state["so_type_first"] = _norm_text(row[col_idx["SO_Type"]])

                # Build sheet-level output
                out_rows = []
                for rnn, s in aggs.items():
                    out_rows.append({
                        "repair_no_clean": rnn,
                        "parts_source_year": year_val,
                        "ordered_via_dms": 1,
                        "parts_line_count": s["line_count"],
                        "parts_order_qty_sum": round(s["order_qty_sum"], 4),
                        "parts_no_count": s["parts_no_count"],
                        "parts_ship_count": s["ship_count"],
                        "parts_arrival_count": s["arrival_count"],
                        "parts_order_ts_min": s["order_ts_min"],
                        "parts_pick_ts_min": s["pick_ts_min"],
                        "parts_ship_ts_min": s["ship_ts_min"],
                        "parts_arrival_dt_min": s["arrival_dt_min"],
                        "parts_arrival_dt_max": s["arrival_dt_max"],
                        "parts_shipping_method_first": s["shipping_method_first"],
                        "parts_prodL2_first": s["prodL2_first"],
                        "parts_so_type_first": s["so_type_first"],
                        "parts_truncation_flag": s["truncation_flag"],
                    })

                out = pd.DataFrame(out_rows)

                if not out.empty:
                    out["parts_multi_line_flag"] = (out["parts_line_count"] > 1).astype("Int8")
                    out["parts_has_shipment_flag"] = (out["parts_ship_count"] > 0).astype("Int8")
                    out["parts_has_arrival_flag"] = (out["parts_arrival_count"] > 0).astype("Int8")

                    out["parts_order_to_pick_days_eda"] = _days_between_cols(
                        out, "parts_order_ts_min", "parts_pick_ts_min")
                    out["parts_pick_to_ship_days_eda"] = _days_between_cols(
                        out, "parts_pick_ts_min", "parts_ship_ts_min")
                    out["parts_ship_to_arrival_days_eda"] = _days_between_cols(
                        out, "parts_ship_ts_min", "parts_arrival_dt_min")
                    out["parts_order_to_arrival_days_eda"] = _days_between_cols(
                        out, "parts_order_ts_min", "parts_arrival_dt_max")

                    out = _clean_leg_durations(out)

                truncated_sheet = (total + 1) >= cfg.excel_max_rows
                yearly_path = cfg.output_dir / f"parts_{year_val}_{fp.stem}.parquet"
                sanitize_for_parquet(out).to_parquet(yearly_path, index=False)
                yearly_paths.append(yearly_path)

                logger.info(
                    "  %s: %s / %s | data rows: %s | unique repairs: %s",
                    "⚠ TRUNCATED" if truncated_sheet else "✓ Complete",
                    fp.name, sname, f"{total:,}", f"{len(aggs):,}",
                )
                del out, aggs
                gc.collect()
        finally:
            wb.close()
            gc.collect()

    # Re-aggregate yearly files (handles repairs spanning multiple years)
    parts_concat = pd.concat(
        [pd.read_parquet(p) for p in yearly_paths],
        ignore_index=True
    )

    parts_features = (
        parts_concat
        .groupby("repair_no_clean", as_index=False)
        .agg(
            ordered_via_dms=("ordered_via_dms", "max"),
            parts_line_count=("parts_line_count", "sum"),
            parts_order_qty_sum=("parts_order_qty_sum", "sum"),
            parts_no_count=("parts_no_count", "sum"),
            parts_ship_count=("parts_ship_count", "sum"),
            parts_arrival_count=("parts_arrival_count", "sum"),
            parts_order_ts_min=("parts_order_ts_min", "min"),
            parts_pick_ts_min=("parts_pick_ts_min", "min"),
            parts_ship_ts_min=("parts_ship_ts_min", "min"),
            parts_arrival_dt_min=("parts_arrival_dt_min", "min"),
            parts_arrival_dt_max=("parts_arrival_dt_max", "max"),
            parts_shipping_method_first=("parts_shipping_method_first", _first_non_null),
            parts_prodL2_first=("parts_prodL2_first", _first_non_null),
            parts_so_type_first=("parts_so_type_first", _first_non_null),
            parts_truncation_flag=("parts_truncation_flag", "max"),
        )
    )

    parts_features["parts_multi_line_flag"] = (
        parts_features["parts_line_count"] > 1).astype("Int8")
    parts_features["parts_has_shipment_flag"] = (
        parts_features["parts_ship_count"] > 0).astype("Int8")
    parts_features["parts_has_arrival_flag"] = (
        parts_features["parts_arrival_count"] > 0).astype("Int8")

    parts_features["parts_order_to_pick_days_eda"] = _days_between_cols(
        parts_features, "parts_order_ts_min", "parts_pick_ts_min")
    parts_features["parts_pick_to_ship_days_eda"] = _days_between_cols(
        parts_features, "parts_pick_ts_min", "parts_ship_ts_min")
    parts_features["parts_ship_to_arrival_days_eda"] = _days_between_cols(
        parts_features, "parts_ship_ts_min", "parts_arrival_dt_min")
    parts_features["parts_order_to_arrival_days_eda"] = _days_between_cols(
        parts_features, "parts_order_ts_min", "parts_arrival_dt_max")

    parts_features = _clean_leg_durations(parts_features)

    parts_clean = sanitize_for_parquet(parts_features)
    parts_clean.to_parquet(cfg.output_dir / "parts_features.parquet", index=False)
    parts_clean[PARTS_MODEL_SAFE_COLS].to_parquet(
        cfg.output_dir / "parts_features_modelsafe.parquet", index=False)

    logger.info("STEP 4B — saved parts_features.parquet, shape: %s", parts_clean.shape)
    return parts_clean


# =====================================================================
# SECTION 7: STEP 4C — INTEGRATED MASTER TABLE
# =====================================================================

def _load_master_with_cohort(cfg: IngestionConfig) -> pd.DataFrame:
    """Load every sheet of the master file and apply cohort flags inline.

    Differs from Step 3 by also applying the division and center-type
    exclusion lists (NON_APPLIANCE_DIVISIONS, NON_RTAT_CENTER_TYPES) at
    cohort-flag time. Also derives month features and the is_holdout
    indicator (1 = 2026).
    """
    master_file = next(
        f for f in cfg.data_dir.glob("*.xlsx")
        if cfg.master_pattern in f.name.lower()
    )

    xls = pd.ExcelFile(master_file, engine="openpyxl")
    frames = []

    for sheet in xls.sheet_names:
        year = int(m.group(1)) if (m := re.search(r"(20\d{2})", sheet)) else None
        header = pd.read_excel(master_file, sheet_name=sheet, nrows=0,
                               dtype=object, engine="openpyxl")
        lookup = {normalize_col(c).lower(): normalize_col(c) for c in header.columns}
        usecols = [lookup[w.lower()] for w in MASTER_COLS if w.lower() in lookup]

        df = pd.read_excel(master_file, sheet_name=sheet, usecols=usecols,
                           dtype=object, engine="openpyxl")
        df.columns = [normalize_col(c) for c in df.columns]
        rename = {normalize_col(lookup[w.lower()]): w
                  for w in MASTER_COLS if w.lower() in lookup}
        df = df.rename(columns=rename)
        for w in MASTER_COLS:
            if w not in df.columns:
                df[w] = pd.NA

        df["source_year"] = year
        df["repair_no_clean"] = df["Repair_No"].apply(clean_key)
        df["rtat_days"] = to_numeric_safe(df["RTAT_Numerator"]).astype("float32")
        df["closed_dt"] = to_datetime_safe(df["Warranty_Closed_Date"])
        df["month_num"] = to_numeric_safe(df["Month_"])
        df["month_of_year"] = (df["month_num"] % 100).astype("Int8")
        df["is_weekend_close"] = df["closed_dt"].dt.dayofweek.isin([5, 6]).astype("Int8")

        # Cohort base flags
        df["flag_valid_id"] = df["repair_no_clean"].notna()
        df["flag_valid_rtat"] = df["rtat_days"].notna()
        df["flag_nonneg"] = df["rtat_days"] >= 0
        df["flag_within_max"] = df["rtat_days"] <= 365

        # Exclusion flags
        df["flag_excluded_division"] = df["Division_Name"].isin(
            NON_APPLIANCE_DIVISIONS).astype("Int8")
        df["flag_excluded_center_type"] = df["SVC_Center_Type"].isin(
            NON_RTAT_CENTER_TYPES).astype("Int8")

        # Final cohort
        df["flag_cohort"] = (
            df["flag_valid_id"]
            & df["flag_valid_rtat"]
            & df["flag_nonneg"]
            & df["flag_within_max"]
            & (df["flag_excluded_division"] == 0)
            & (df["flag_excluded_center_type"] == 0)
        )

        df["target_days"] = np.where(
            df["flag_cohort"], df["rtat_days"], np.nan
        ).astype("float32")

        for t in cfg.ontime_thresholds:
            df[f"OnTime_{t}"] = pd.array(
                np.where(df["flag_cohort"],
                         (df["target_days"] <= t).astype("int8"), pd.NA),
                dtype="Int8"
            )

        # NPS flags
        nps_sum = (to_numeric_safe(df["Promoter"]).fillna(0)
                   + to_numeric_safe(df["Passive"]).fillna(0)
                   + to_numeric_safe(df["Detractor"]).fillna(0))
        df["has_nps"] = (nps_sum > 0).astype("Int8")
        df["is_promoter"] = to_numeric_safe(df["Promoter"]).gt(0).astype("Int8")
        df["is_detractor"] = to_numeric_safe(df["Detractor"]).gt(0).astype("Int8")

        # Holdout flag
        df["is_holdout"] = (df["source_year"] == 2026).astype("Int8")

        frames.append(df)
        excl = int((df["flag_excluded_division"] | df["flag_excluded_center_type"]).sum())
        logger.info(
            "  %s: %s rows | cohort: %s | excluded: %s",
            sheet, f"{len(df):,}",
            f"{int(df['flag_cohort'].sum()):,}",
            f"{excl:,}",
        )
        gc.collect()

    master = pd.concat(frames, ignore_index=True)
    del frames
    gc.collect()
    return master


def _derive_parts_safe_feature(master: pd.DataFrame) -> pd.DataFrame:
    """Compute the leakage-safe parts delivery duration feature.

    The raw ``parts_order_to_arrival_days_eda`` includes deliveries that
    arrived *after* the repair was closed — an impossible operational
    signal that leaks future information into the model. The "safe"
    variant nulls out durations where arrival did not precede close.

    The Warranty_Closed_Date is stored as YYYYMMDD integers in the
    source data and requires the explicit ``format='%Y%m%d'`` argument
    on pd.to_datetime — generic parsing fails on these.
    """
    arrival_dt = pd.to_datetime(master["parts_arrival_dt_max"], errors="coerce")
    closed_dt = pd.to_datetime(
        master["Warranty_Closed_Date"].astype(str).str.strip(),
        format="%Y%m%d",
        errors="coerce",
    )

    # Strip timezone if present
    if hasattr(arrival_dt.dt, "tz") and arrival_dt.dt.tz is not None:
        arrival_dt = arrival_dt.dt.tz_localize(None)
    if hasattr(closed_dt.dt, "tz") and closed_dt.dt.tz is not None:
        closed_dt = closed_dt.dt.tz_localize(None)

    master["parts_arrive_before_close"] = (
        arrival_dt.notna() & closed_dt.notna() & (arrival_dt < closed_dt)
    ).astype("Int8")

    master["parts_order_to_arrival_days_safe"] = np.where(
        master["parts_arrive_before_close"] == 1,
        master["parts_order_to_arrival_days_eda"],
        np.nan,
    ).astype("float32")
    return master


def run_step4c_integrate(cfg: IngestionConfig) -> dict[str, pd.DataFrame]:
    """Join master + reclaim + parts into a single integrated table.

    Steps:
        1. Load master with cohort flags (uses exclusion lists)
        2. Read reclaim features parquet, rename cols to avoid join clash
        3. Left-join reclaim onto master
        4. Read parts features parquet, left-join onto master
        5. Fill DMS non-match as 0, use reclaim parts flag as ground truth
        6. Compute parts_shipping_tier from raw shipping method
        7. Compute engineer_hist_mean_rtat on training years only
        8. Compute leakage-safe parts delivery duration
        9. Split into train (2023-2025 cohort) and holdout (2026 cohort)

    Persists three parquets:
        master_integrated.parquet — full universe
        master_train.parquet      — training cohort
        master_holdout.parquet    — holdout cohort
    """
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    # 4C-1: Load master with cohort
    logger.info("STEP 4C — loading master with cohort flags")
    master = _load_master_with_cohort(cfg)
    logger.info("Master loaded: %s rows", f"{len(master):,}")

    # 4C-3: Join reclaim
    logger.info("Joining reclaim features...")
    reclaim_features = pd.read_parquet(cfg.output_dir / "reclaim_features.parquet")
    reclaim_features = reclaim_features.rename(columns={
        "source_year": "reclaim_source_year",
        "Division_Code": "rcl_Division_Code",
        "Division_Name": "rcl_Division_Name",
        "Product2_Name": "rcl_Product2_Name",
        "Product3_Name": "rcl_Product3_Name",
        "Model_Code": "rcl_Model_Code",
        "Warranty_Flag": "rcl_Warranty_Flag",
    })
    pre = len(master)
    master = master.merge(reclaim_features, on="repair_no_clean", how="left")
    logger.info("  Rows before/after: %s / %s | reclaim match: %.1f%%",
                f"{pre:,}", f"{len(master):,}",
                100 * master["has_parts_reclaim"].notna().sum() / len(master))

    # 4C-4: Join parts
    logger.info("Joining parts features...")
    parts_features = pd.read_parquet(cfg.output_dir / "parts_features.parquet")
    pre = len(master)
    master = master.merge(parts_features, on="repair_no_clean", how="left")
    logger.info("  Rows before/after: %s / %s | DMS match: %.1f%%",
                f"{pre:,}", f"{len(master):,}",
                100 * master["ordered_via_dms"].notna().sum() / len(master))

    # Fill DMS non-match as 0
    master["ordered_via_dms"] = master["ordered_via_dms"].fillna(0).astype("Int8")

    # Parts-required ground truth: reclaim first, DMS as fallback
    master["has_parts_reclaim"] = master["has_parts_reclaim"].fillna(
        master["ordered_via_dms"]
    ).astype("Int8")

    # Shipping tier
    master["parts_shipping_tier"] = master["parts_shipping_method_first"].apply(
        simplify_shipping
    )

    # 4C-5: Engineering proxies
    logger.info("Engineering integrated features...")
    train_mask = master["is_holdout"] == 0
    eng_mean = (
        master[train_mask & master["flag_cohort"]]
        .groupby("SVC_Engineer_Code")["target_days"]
        .mean()
        .rename("engineer_hist_mean_rtat")
    )
    master = master.merge(eng_mean, on="SVC_Engineer_Code", how="left")

    master["parts_count_reclaim"] = master["parts_count_reclaim"].fillna(0).astype("Int8")

    # 4C-5 continued: leakage-safe parts duration
    master = _derive_parts_safe_feature(master)

    # 4C-6: Save splits
    master_clean = sanitize_for_parquet(master)
    train = master_clean[(master_clean["is_holdout"] == 0)
                         & (master_clean["flag_cohort"])]
    holdout = master_clean[(master_clean["is_holdout"] == 1)
                           & (master_clean["flag_cohort"])]

    master_clean.to_parquet(cfg.output_dir / "master_integrated.parquet", index=False)
    train.to_parquet(cfg.output_dir / "master_train.parquet", index=False)
    holdout.to_parquet(cfg.output_dir / "master_holdout.parquet", index=False)

    logger.info("STEP 4C — integrated: %s | train: %s | holdout: %s",
                f"{len(master_clean):,}", f"{len(train):,}", f"{len(holdout):,}")

    return {
        "master_integrated": master_clean,
        "master_train": train,
        "master_holdout": holdout,
    }


# =====================================================================
# MAIN ORCHESTRATION
# =====================================================================

def run_ingestion(cfg: IngestionConfig | None = None) -> dict:
    """Run the full ingestion pipeline (Steps 1 → 4C) end-to-end.

    Returns:
        Dictionary of stage names → DataFrames produced. All artifacts
        are also persisted to ``cfg.output_dir`` for downstream stages.
    """
    cfg = cfg or IngestionConfig()

    logger.info("=== Step 1: Workbook inventory ===")
    step1 = run_step1_inventory(cfg)

    logger.info("=== Step 2: Key validation ===")
    step2 = run_step2_key_validation(cfg)

    logger.info("=== Step 3: Cohort + targets ===")
    step3 = run_step3_cohort(cfg)

    logger.info("=== Step 4A: Reclaim features ===")
    reclaim = run_step4a_reclaim(cfg)

    logger.info("=== Step 4B: Parts features ===")
    parts = run_step4b_parts(cfg)

    logger.info("=== Step 4C: Integrated master ===")
    step4c = run_step4c_integrate(cfg)

    return {
        **step1, **step2, **step3,
        "reclaim_features": reclaim,
        "parts_features": parts,
        **step4c,
    }


def _configure_logging() -> None:
    """Configure root logging for CLI execution."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )


if __name__ == "__main__":
    _configure_logging()
    artifacts = run_ingestion()
    logger.info("Ingestion complete. %d artifacts produced.", len(artifacts))
